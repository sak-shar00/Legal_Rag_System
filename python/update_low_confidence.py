"""
update_low_confidence.py

Re-tests only the "Low" confidence rows in your golden set against your
LIVE running API (the one with real vector+BM25 hybrid search + Groq LLM),
and writes the improved results into a new column set for you to quickly
review and confirm.

HOW TO USE:
1. Make sure your FastAPI server is running (uvicorn ... on port 8000)
2. Put this file anywhere, along with golden_set_UPDATED.xlsx
3. pip install requests openpyxl
4. Run: python update_low_confidence.py
5. Open golden_set_REVIEWED.xlsx - only Low-confidence rows changed,
   High/Medium rows are left untouched.
"""

import requests
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

API_URL = "http://127.0.0.1:8000/ask"
# IMPORTANT: After your first run, change this to "golden_set_REVIEWED.xlsx"
# so already-updated rows (marked "REVIEW") are skipped and only the
# remaining "Low" rows get retried. Don't re-run against the original
# golden_set_UPDATED.xlsx every time, or you'll redo completed work.
INPUT_FILE = "golden_set_REVIEWED.xlsx"
OUTPUT_FILE = "golden_set_REVIEWED.xlsx"

print(f"Loading {INPUT_FILE}...")
wb = load_workbook(INPUT_FILE)
sheet = wb.active

headers = [c.value for c in sheet[1]]
print("Columns:", headers)

# Adjust these if your column names differ
col_query = headers.index("Query") + 1
col_answer = headers.index("Ground Truth Answer") + 1
col_conf = headers.index("Confidence") + 1
col_doc = headers.index("Best Source Document") + 1
col_pages = headers.index("Page Number") + 1
col_snippet = headers.index("Retrieved Snippet (for reference)") + 1

updated_count = 0
skipped_count = 0

for row_idx in range(2, sheet.max_row + 1):
    confidence = sheet.cell(row=row_idx, column=col_conf).value
    query = sheet.cell(row=row_idx, column=col_query).value

    if confidence != "Low":
        skipped_count += 1
        continue

    print(f"\n[Row {row_idx}] Re-testing: {query[:70]}...")

    try:
        resp = requests.post(API_URL, json={"question": query}, timeout=60)
        resp.raise_for_status()
        result = resp.json()
    except Exception as e:
        print(f"  ERROR calling API (HTTP level): {e}")
        time.sleep(3)
        continue

    if not result.get("success", True) or not result.get("sources"):
        # Print the ACTUAL reason instead of just guessing "no sources"
        reason = result.get("answer", "unknown reason")
        print(f"  FAILED - success={result.get('success')} | reason: {reason[:200]}")
        if not result.get("sources"):
            sheet.cell(row=row_idx, column=col_answer).value = (
                "[NO GOOD MATCH FOUND EVEN WITH LIVE HYBRID SEARCH - "
                "this query may genuinely not be covered by the 100 documents]"
            )
        time.sleep(3)
        continue

    top_source = result["sources"][0]
    new_answer = result["answer"]

    # Write updated values
    sheet.cell(row=row_idx, column=col_answer).value = new_answer
    sheet.cell(row=row_idx, column=col_doc).value = top_source.get("document", top_source.get("filename", ""))
    sheet.cell(row=row_idx, column=col_pages).value = f"{top_source.get('page_start','?')}-{top_source.get('page_end','?')}"
    sheet.cell(row=row_idx, column=col_conf).value = "REVIEW"  # mark for you to eyeball, then change to High/Medium

    # Highlight the row so it's easy to spot what changed
    fill = PatternFill("solid", start_color="D9E8FB", end_color="D9E8FB")
    for c in range(1, len(headers) + 1):
        sheet.cell(row=row_idx, column=c).fill = fill

    updated_count += 1
    print(f"  Updated. New top source: {top_source.get('document', top_source.get('filename',''))}")

    time.sleep(3)  # increased delay to avoid Groq rate limiting

wb.save(OUTPUT_FILE)
print(f"\nDone! Updated {updated_count} rows, skipped {skipped_count} (already High/Medium).")
print(f"Saved to {OUTPUT_FILE}")
print("\nNext: open the file, check the blue-highlighted rows (marked 'REVIEW'),")
print("read the new answer, and change Confidence to High/Medium/Low as appropriate.")